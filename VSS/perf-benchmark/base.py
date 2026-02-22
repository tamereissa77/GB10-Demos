######################################################################################################
# SPDX-FileCopyrightText: Copyright (c) 2025 NVIDIA CORPORATION & AFFILIATES. All rights reserved.
# SPDX-License-Identifier: Apache-2.0
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
######################################################################################################
"""
Base class for VSS performance benchmarks.
"""

import json
import logging
import os
import shutil
import threading
from abc import ABC, abstractmethod
from typing import Any, Dict, List, Optional

import pandas as pd
import requests
import yaml
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from gpu_monitor import GPUMonitor
except ImportError:
    GPUMonitor = None


class BenchmarkBase(ABC):
    """
    Base class for all VSS performance benchmarks.

    Provides common infrastructure for:
    - Configuration parsing and validation
    - API communication with retry logic
    - GPU monitoring lifecycle management
    - Directory and file management
    - Results collection and Excel generation
    """

    # HTTP client defaults
    DEFAULT_RETRY_COUNT = 3  # Number of retry attempts
    DEFAULT_RETRY_BACKOFF_FACTOR = 1  # Exponential backoff factor
    DEFAULT_RETRY_STATUS_CODES = [429, 500, 502, 503, 504]  # Status codes to retry
    DEFAULT_POOL_CONNECTIONS = 1024  # Connection pool size
    DEFAULT_POOL_MAXSIZE = 1024  # Maximum pool size
    DEFAULT_THREAD_WAIT_TIMEOUT = 30  # Thread completion timeout

    def __init__(self, base_url: str, output_base_dir: str = "vss-perf-report"):
        """
        Initialize the benchmark.

        Args:
            base_url: Base URL for the VSS API
            output_base_dir: Base directory for all benchmark outputs
        """
        self.base_url = base_url.rstrip("/")
        self.output_base_dir = output_base_dir

        # Setup HTTP session with connection pooling and retries
        self.session = requests.Session()
        self._configure_http_session()

        # GPU monitoring setup (will be configured in parse_config)
        self.gpu_monitor = None
        self.gpu_monitoring_config = {}
        self.vlm_gpus = []
        self.llm_gpus = []

        # Setup logging
        self._setup_logging()

        # Cleanup tracking
        self.active_resources = []
        self.cleanup_lock = threading.Lock()

        # Current execution state
        self.current_scenario_dir = None
        self.gpu_monitoring_active = False

    def _setup_logging(self):
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s - %(levelname)s - %(message)s",
            handlers=[logging.StreamHandler()],
        )
        self.logger = logging.getLogger(self.__class__.__name__)

    def _configure_http_session(self, pool_maxsize=None):
        """Configure HTTP session with connection pooling and retry logic"""
        # Use provided pool_maxsize or default
        pool_size = pool_maxsize if pool_maxsize is not None else self.DEFAULT_POOL_MAXSIZE

        retry_strategy = Retry(
            total=self.DEFAULT_RETRY_COUNT,
            backoff_factor=self.DEFAULT_RETRY_BACKOFF_FACTOR,
            status_forcelist=self.DEFAULT_RETRY_STATUS_CODES,
        )

        adapter = HTTPAdapter(
            pool_connections=self.DEFAULT_POOL_CONNECTIONS,
            pool_maxsize=pool_size,
            max_retries=retry_strategy,
        )

        self.session.mount("http://", adapter)
        self.session.mount("https://", adapter)

        if pool_maxsize is not None:
            self.logger.debug(f"Reconfigured HTTP session with pool_maxsize={pool_size}")

    # ================================
    # ABSTRACT METHODS - Must implement
    # ================================

    @abstractmethod
    def parse_benchmark_config(self, scenario_config: Dict, global_config: Dict) -> Any:
        """
        Parse and validate benchmark-specific configuration.

        Args:
            scenario_config: Scenario-specific configuration
            global_config: Global configuration settings

        Returns:
            Parsed benchmark configuration object

        Raises:
            ValueError: If configuration is invalid
        """
        pass

    @abstractmethod
    def execute(self, config: Dict, scenario_name: str) -> Dict[str, Any]:
        """
        Execute the benchmark and return results metadata.

        Args:
            config: Full configuration dictionary
            scenario_name: Name of the scenario being executed

        Returns:
            Dictionary containing execution results and metadata

        Raises:
            Exception: If benchmark execution fails
        """
        pass

    @abstractmethod
    def analyze_results(self, results_dir: str, output_file: str) -> None:
        """
        Generate Excel report from benchmark results.

        Args:
            results_dir: Directory containing benchmark results
            output_file: Path to output Excel file

        Raises:
            Exception: If analysis or report generation fails
        """
        pass

    # ================================
    # CONFIGURATION METHODS
    # ================================

    def _load_defaults(self) -> Dict[str, Any]:
        """Load default parameters from defaults.yaml"""
        defaults_path = os.path.join(os.path.dirname(__file__), "defaults.yaml")
        if not os.path.exists(defaults_path):
            raise FileNotFoundError(
                f"defaults.yaml not found at {defaults_path}. "
                "This file is required for benchmark configuration."
            )
        with open(defaults_path, "r") as f:
            defaults = yaml.safe_load(f)
            if not defaults:
                raise ValueError("defaults.yaml is empty or invalid")
            return defaults

    def _merge_with_defaults(self, user_params: Dict, default_params: Dict) -> Dict:
        """Merge user parameters with defaults"""
        merged = default_params.copy() if default_params else {}
        if user_params:
            merged.update(user_params)
        return merged

    def parse_global_config(self, config: Dict) -> Dict[str, Any]:
        """
        Parse and validate global configuration.

        Args:
            config: Full configuration dictionary

        Returns:
            Validated global configuration

        Raises:
            ValueError: If global configuration is invalid
        """
        if "global" not in config:
            raise ValueError("Missing 'global' section in configuration")

        global_config = config["global"]
        defaults = self._load_defaults()

        # Validate required fields
        required_fields = ["vlm_gpus", "llm_gpus"]
        for field in required_fields:
            if field not in global_config:
                raise ValueError(f"Missing required global config field: {field}")

        # Merge each API parameter section with defaults
        global_config["summarize_api_params"] = self._merge_with_defaults(
            global_config.get("summarize_api_params", {}), defaults["summarize_api_params"]
        )
        global_config["chat_api_params"] = self._merge_with_defaults(
            global_config.get("chat_api_params", {}), defaults["chat_api_params"]
        )
        global_config["vlm_captions_params"] = self._merge_with_defaults(
            global_config.get("vlm_captions_params", {}), defaults["vlm_captions_params"]
        )
        global_config["alert_review_params"] = self._merge_with_defaults(
            global_config.get("alert_review_params", {}), defaults["alert_review_params"]
        )

        # Setup GPU configuration
        self.vlm_gpus = global_config["vlm_gpus"]
        self.llm_gpus = global_config["llm_gpus"]
        self.gpu_monitoring_config = global_config.get("gpu_monitoring", {"enabled": False})

        # Initialize GPU monitor if enabled
        if self.gpu_monitoring_config.get("enabled", False) and GPUMonitor is not None:
            all_gpu_ids = list(set(self.vlm_gpus + self.llm_gpus))
            if all_gpu_ids:
                self.gpu_monitor = GPUMonitor(gpu_ids=all_gpu_ids)
                self.logger.debug(f"GPU monitoring enabled for GPUs: {all_gpu_ids}")
            else:
                self.logger.warning("GPU monitoring enabled but no GPU IDs specified")
        else:
            self.logger.debug("GPU monitoring disabled")

        return global_config

    def get_vlm_dimensions(self, scenario_config: Dict, global_config: Dict) -> Dict[str, int]:
        """
        Get VLM dimensions with scenario override support.

        Args:
            scenario_config: Scenario-specific configuration
            global_config: Global configuration

        Returns:
            Dictionary with vlm_input_width and vlm_input_height
        """
        # Get VLM dimensions with scenario override logic
        global_vlm_dimensions = global_config.get("vlm_dimensions", {})
        scenario_vlm_dimensions = scenario_config.get("vlm_dimensions", {})

        # Use scenario dimensions if available, otherwise use global defaults
        vlm_input_width = scenario_vlm_dimensions.get(
            "input_width", global_vlm_dimensions.get("input_width", 0)
        )
        vlm_input_height = scenario_vlm_dimensions.get(
            "input_height", global_vlm_dimensions.get("input_height", 0)
        )

        return {"vlm_input_width": vlm_input_width, "vlm_input_height": vlm_input_height}

    def load_config_file(self, config_path: str) -> Dict[str, Any]:
        """
        Load and parse YAML configuration file.

        Args:
            config_path: Path to YAML configuration file

        Returns:
            Parsed configuration dictionary

        Raises:
            FileNotFoundError: If config file doesn't exist
            ValueError: If YAML is invalid
        """
        try:
            with open(config_path, "r") as f:
                config = yaml.safe_load(f)

            # Basic structure validation
            if not isinstance(config, dict):
                raise ValueError("Configuration must be a dictionary")

            if "global" not in config:
                raise ValueError("Missing 'global' section")

            if "test_scenarios" not in config:
                raise ValueError("Missing 'test_scenarios' section")

            return config

        except FileNotFoundError:
            raise FileNotFoundError(f"Configuration file not found: {config_path}")
        except yaml.YAMLError as e:
            raise ValueError(f"Invalid YAML in configuration file: {e}")

    # ================================
    # EXECUTION INFRASTRUCTURE
    # ================================

    def setup_scenario_directory(self, scenario_name: str) -> str:
        """
        Create and setup directory for scenario execution.

        Args:
            scenario_name: Name of the scenario

        Returns:
            Path to created scenario directory
        """
        scenario_dir = os.path.join(self.output_base_dir, scenario_name)

        # Remove existing directory if it exists
        if os.path.exists(scenario_dir):
            shutil.rmtree(scenario_dir)

        # Create new directory
        os.makedirs(scenario_dir, exist_ok=True)
        self.current_scenario_dir = scenario_dir

        self.logger.debug(f"Created scenario directory: {scenario_dir}")
        return scenario_dir

    def start_gpu_monitoring(self):
        """Start GPU monitoring if configured"""
        if self.gpu_monitor and not self.gpu_monitoring_active:
            try:
                interval = self.gpu_monitoring_config.get("interval_seconds", 2)
                self.gpu_monitor.start_recording_gpu_usage(interval_in_seconds=interval)
                self.gpu_monitor.start_recording_nvdec(interval_in_seconds=interval)
                self.gpu_monitoring_active = True
                self.logger.debug("GPU monitoring started")
            except Exception as e:
                self.logger.error(f"Failed to start GPU monitoring: {e}")

    def stop_gpu_monitoring(self, export_dir: str = None, filename_prefix: str = "gpu_metrics"):
        """
        Stop GPU monitoring and optionally export data.

        Args:
            export_dir: Directory to export GPU data (optional)
            filename_prefix: Prefix for exported files
        """
        if self.gpu_monitor and self.gpu_monitoring_active:
            try:
                self.gpu_monitor.stop_recording_gpu()
                self.gpu_monitor.stop_recording_nvdec()
                self.gpu_monitoring_active = False

                # Export data if directory specified
                if export_dir:
                    export_csv = self.gpu_monitoring_config.get("export_csv", False)
                    export_plots = self.gpu_monitoring_config.get("export_plots", False)
                    self.gpu_monitor.export_data(
                        output_dir=export_dir,
                        base_filename=filename_prefix,
                        export_csv=export_csv,
                        export_plots=export_plots,
                    )
                    self.logger.debug(f"GPU data exported to: {export_dir}")

                self.logger.debug("GPU monitoring stopped")
            except Exception as e:
                self.logger.error(f"Failed to stop GPU monitoring: {e}")

    # ================================
    # API COMMUNICATION
    # ================================

    def make_api_call(
        self,
        endpoint: str,
        method: str = "GET",
        data: Optional[Dict] = None,
        params: Optional[Dict] = None,
        headers: Optional[Dict] = None,
        files: Optional[Dict] = None,
    ) -> requests.Response:
        """
        Make an API call to the VSS backend.

        Args:
            endpoint: API endpoint (will be appended to base_url)
            method: HTTP method
            data: JSON data for request body
            params: URL parameters
            headers: HTTP headers
            files: Files for multipart upload

        Returns:
            Response object

        Raises:
            requests.RequestException: If API call fails
        """
        url = f"{self.base_url}/{endpoint.lstrip('/')}"

        request_kwargs = {"method": method, "url": url, "params": params, "headers": headers}

        if data and method in ["POST", "PUT", "PATCH"]:
            request_kwargs["json"] = data

        if files:
            request_kwargs["files"] = files

        try:
            self.logger.debug(f"API call: {method} {url}")
            response = self.session.request(**request_kwargs)
            response.raise_for_status()

            if not self.validate_api_response(response):
                raise requests.exceptions.RequestException(f"Invalid API response from {endpoint}")

            return response
        except requests.exceptions.RequestException as e:
            self.logger.error(f"API call failed: {e}")
            raise

    def validate_api_response(self, response: requests.Response) -> bool:
        """
        Validate API response for common issues.

        Args:
            response: Response to validate

        Returns:
            True if response is valid, False otherwise
        """
        # Check status code
        if response.status_code >= 400:
            try:
                error_data = response.json()
                self.logger.error(f"API error: {error_data}")
            except Exception:
                self.logger.error(f"API error: HTTP {response.status_code}")
            return False

        if not response.content and response.status_code >= 300:
            self.logger.error("Empty API response for non-2xx status")
            return False

        # Validate JSON if content-type suggests it and there's content
        if response.content and "application/json" in response.headers.get("Content-Type", ""):
            try:
                response.json()
            except json.JSONDecodeError:
                self.logger.error("Invalid JSON in API response")
                return False

        return True

    def get_available_models(self) -> str:
        """Get available model from the API"""
        response = self.make_api_call("/models")
        data = response.json()
        return data["data"][0]["id"]

    def scrape_metrics(self) -> Dict[str, Any]:
        """
        Scrape metrics from the /metrics endpoint and parse them into a dictionary.

        Returns:
            Dictionary of metrics with numeric values
        """
        try:
            response = self.make_api_call("/metrics")
            metrics = {}
            for line in response.text.split("\n"):
                if line and not line.startswith("#"):
                    try:
                        name, value = line.split(" ")
                        if any(name.endswith(suffix) for suffix in ["_latest", "_sum", "_count"]):
                            metrics[name] = float(value)
                    except ValueError:
                        continue
            return metrics
        except Exception as e:
            self.logger.error(f"Failed to scrape metrics: {e}")
            return {}

    # ================================
    # GPU STATISTICS PROCESSING
    # ================================

    def process_gpu_stats(self, gpu_stats_file: str) -> Dict[str, Any]:
        """
        Process GPU statistics from exported JSON file.

        Args:
            gpu_stats_file: Path to GPU stats JSON file

        Returns:
            Dictionary with processed GPU metrics
        """
        if not os.path.exists(gpu_stats_file):
            self.logger.warning(f"GPU stats file not found: {gpu_stats_file}")
            return {}

        try:
            with open(gpu_stats_file, "r") as f:
                gpu_data = json.load(f)

            gpu_metrics = {}

            # Process VLM GPU stats
            vlm_stats = []
            for gpu_id in self.vlm_gpus:
                gpu_key = f"GPU{gpu_id}"
                if gpu_key in gpu_data.get("gpu_stats", {}):
                    vlm_stats.append(gpu_data["gpu_stats"][gpu_key])

            if vlm_stats:
                vlm_usage = [stats.get("usage", {}).get("mean", 0) for stats in vlm_stats]
                vlm_memory = [stats.get("memory", {}).get("mean", 0) for stats in vlm_stats]
                vlm_usage_p90 = [stats.get("usage", {}).get("p90", 0) for stats in vlm_stats]

                gpu_metrics.update(
                    {
                        "vlm_gpu_usage_mean": sum(vlm_usage) / len(vlm_usage) if vlm_usage else 0,
                        "vlm_gpu_memory_mean": (
                            sum(vlm_memory) / len(vlm_memory) if vlm_memory else 0
                        ),
                        "vlm_gpu_usage_p90": (
                            sum(vlm_usage_p90) / len(vlm_usage_p90) if vlm_usage_p90 else 0
                        ),
                    }
                )

            # Process LLM GPU stats
            llm_stats = []
            for gpu_id in self.llm_gpus:
                gpu_key = f"GPU{gpu_id}"
                if gpu_key in gpu_data.get("gpu_stats", {}):
                    llm_stats.append(gpu_data["gpu_stats"][gpu_key])

            if llm_stats:
                llm_usage = [stats.get("usage", {}).get("mean", 0) for stats in llm_stats]
                llm_memory = [stats.get("memory", {}).get("mean", 0) for stats in llm_stats]
                llm_usage_p90 = [stats.get("usage", {}).get("p90", 0) for stats in llm_stats]

                gpu_metrics.update(
                    {
                        "llm_gpu_usage_mean": sum(llm_usage) / len(llm_usage) if llm_usage else 0,
                        "llm_gpu_memory_mean": (
                            sum(llm_memory) / len(llm_memory) if llm_memory else 0
                        ),
                        "llm_gpu_usage_p90": (
                            sum(llm_usage_p90) / len(llm_usage_p90) if llm_usage_p90 else 0
                        ),
                    }
                )

            # Process NVDEC stats (typically on VLM GPUs)
            nvdec_stats = gpu_data.get("nvdec_stats", {})
            vlm_nvdec_stats = []
            for gpu_id in self.vlm_gpus:
                nvdec_key = f"GPU{gpu_id}_AvgNVDEC"
                # Try to get individual GPU NVDEC stats if available
                if nvdec_key in gpu_data.get("nvdec_data", {}):
                    vlm_nvdec_stats.extend(gpu_data["nvdec_data"][nvdec_key])

            if vlm_nvdec_stats:
                import numpy as np

                gpu_metrics.update(
                    {
                        "vlm_nvdec_usage_mean": float(np.mean(vlm_nvdec_stats)),
                        "vlm_nvdec_usage_std": float(np.std(vlm_nvdec_stats)),
                        "vlm_nvdec_usage_p90": float(np.percentile(vlm_nvdec_stats, 90)),
                    }
                )
            else:
                # If no individual GPU NVDEC data, use overall NVDEC for VLM
                if nvdec_stats:
                    gpu_metrics.update(
                        {
                            "vlm_nvdec_usage_mean": nvdec_stats.get("mean", 0),
                            "vlm_nvdec_usage_std": nvdec_stats.get("std", 0),
                            "vlm_nvdec_usage_p90": nvdec_stats.get("p90", 0),
                        }
                    )

            return gpu_metrics

        except Exception as e:
            self.logger.error(f"Error processing GPU stats: {e}")
            return {}

    # ================================
    # UTILITY METHODS
    # ================================

    def save_json_data(self, data: Dict, filepath: str):
        """Save dictionary data as JSON file"""
        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            with open(filepath, "w") as f:
                json.dump(data, f, indent=2)
        except Exception as e:
            self.logger.error(f"Failed to save JSON data to {filepath}: {e}")

    def save_response(self, response: requests.Response, test_dir: str, filename: str):
        """Save response data to a file"""
        filepath = os.path.join(test_dir, filename)
        try:
            with open(filepath, "w") as f:
                json.dump(response.json(), f, indent=4)
        except Exception as e:
            self.logger.error(f"Failed to save response: {str(e)}")

    def round_floats(self, obj, decimal_places: int = 2):
        """Recursively round all floating point numbers in a data structure"""
        if isinstance(obj, float):
            return round(obj, decimal_places)
        elif isinstance(obj, dict):
            return {key: self.round_floats(value, decimal_places) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self.round_floats(item, decimal_places) for item in obj]
        elif isinstance(obj, tuple):
            return tuple(self.round_floats(item, decimal_places) for item in obj)
        else:
            return obj

    def add_plots_to_excel(
        self, excel_path: str, df: pd.DataFrame, x_column: str, latency_columns: List[str]
    ):
        """
        Create latency plots using matplotlib and embed them in the Excel workbook.

        Args:
            excel_path: Path to the Excel file
            df: DataFrame with data to plot
            x_column: Column to use for x-axis
            latency_columns: List of columns to plot as y-values
        """
        try:
            import io

            import matplotlib.pyplot as plt
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image

            # Load the existing Excel workbook
            workbook = load_workbook(excel_path)

            # Calculate starting position for images (after the data)
            start_row = len(df) + 5

            # Create latency chart
            plt.figure(figsize=(12, 8))

            # Check if this is file burst mode by looking for benchmark_mode column and concurrency_level data
            is_file_burst = (
                "benchmark_mode" in df.columns
                and "concurrency_level" in df.columns
                and "file_burst" in df["benchmark_mode"].values
            )

            if is_file_burst:
                # For file burst mode: plot each test case as a separate line
                test_cases = (
                    df["test_case_id"].unique() if "test_case_id" in df.columns else [df.index[0]]
                )
                colors = ["red", "green", "blue", "orange", "purple", "brown"]

                for test_idx, test_case in enumerate(test_cases):
                    test_data = (
                        df[df["test_case_id"] == test_case] if "test_case_id" in df.columns else df
                    )

                    for i, col_name in enumerate(latency_columns):
                        if col_name in test_data.columns:
                            x_values = test_data[x_column].sort_values()
                            y_values = test_data.set_index(x_column)[col_name].reindex(x_values)

                            # Create a unique label for each test case and metric combination
                            if len(test_cases) > 1:
                                label = f"{test_case} - {col_name.replace('_', ' ').title()}"
                            else:
                                label = col_name.replace("_", " ").title()

                            plt.plot(
                                x_values,
                                y_values,
                                marker="o",
                                linewidth=2,
                                label=label,
                                color=colors[(test_idx * len(latency_columns) + i) % len(colors)],
                                markersize=6,
                            )

                plt.title("File Burst Mode: Latency vs Test Case", fontsize=14, fontweight="bold")
                plt.xlabel("Test Case ID", fontsize=12)
                plt.ylabel("Latency (seconds)", fontsize=12)
            else:
                # Original logic for single file mode: group by x_column and calculate statistics
                grouped_data = df.groupby(x_column)

                # Plot each latency column with error bars
                colors = ["red", "green", "blue", "orange"]
                for i, col_name in enumerate(latency_columns):
                    if col_name in df.columns:
                        # Calculate mean and std for each group
                        means = grouped_data[col_name].mean()
                        stds = grouped_data[col_name].std()
                        x_values = means.index

                        plt.errorbar(
                            x_values,
                            means,
                            yerr=stds,
                            marker="o",
                            linewidth=3,
                            capsize=10,
                            capthick=3,
                            label=col_name.replace("_", " ").title(),
                            color=colors[i % len(colors)],
                            elinewidth=3,
                            markersize=8,
                        )

                plt.title("Latency Trends (Mean ± Std)", fontsize=14, fontweight="bold")
                plt.xlabel(x_column.replace("_", " ").title(), fontsize=12)
                plt.ylabel("Latency (seconds)", fontsize=12)

            plt.legend(fontsize=10)
            plt.grid(True, alpha=0.3)
            plt.xticks(rotation=45)
            plt.tight_layout()

            # Save plot to bytes buffer
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format="png", dpi=300, bbox_inches="tight")
            img_buffer.seek(0)
            plt.close()

            # Add image to Excel
            img = Image(img_buffer)
            img.width = 600  # Adjust width
            img.height = 400  # Adjust height

            # Add to Summary sheet
            if "Summary" in workbook.sheetnames:
                worksheet = workbook["Summary"]
                worksheet.add_image(img, f"A{start_row}")

            # Save the workbook with embedded images
            workbook.save(excel_path)
            self.logger.debug(f"Plots added to Excel file: {excel_path}")

        except ImportError as e:
            self.logger.warning(f"Plotting libraries not available, skipping plots: {e}")
        except Exception as e:
            self.logger.error(f"Failed to add plots to Excel: {e}")
            # Don't raise - let the Excel file be saved without plots

    def get_gpu_info_dataframe(self) -> pd.DataFrame:
        """
        Get GPU information as a pandas DataFrame for Excel reports

        Returns:
            DataFrame containing GPU specifications
        """
        gpu_info = []

        try:
            import pynvml

            pynvml.nvmlInit()
            device_count = pynvml.nvmlDeviceGetCount()

            for i in range(device_count):
                handle = pynvml.nvmlDeviceGetHandleByIndex(i)

                # Basic info
                name = pynvml.nvmlDeviceGetName(handle)

                # Memory info
                mem_info = pynvml.nvmlDeviceGetMemoryInfo(handle)
                total_mem_gb = mem_info.total / (1024**3)

                # Driver version (only need once)
                driver_version = pynvml.nvmlSystemGetDriverVersion() if i == 0 else ""

                # CUDA compute capability
                major, minor = pynvml.nvmlDeviceGetCudaComputeCapability(handle)
                compute_capability = f"{major}.{minor}"

                # Clock speeds
                try:
                    max_gpu_clock = pynvml.nvmlDeviceGetMaxClockInfo(
                        handle, pynvml.NVML_CLOCK_GRAPHICS
                    )
                    max_mem_clock = pynvml.nvmlDeviceGetMaxClockInfo(handle, pynvml.NVML_CLOCK_MEM)
                except Exception:
                    max_gpu_clock = "N/A"
                    max_mem_clock = "N/A"

                gpu_info.append(
                    {
                        "GPU Index": i,
                        "Name": name,
                        "Total Memory (GB)": f"{total_mem_gb:.2f}",
                        "Compute Capability": compute_capability,
                        "Max GPU Clock (MHz)": max_gpu_clock,
                        "Max Memory Clock (MHz)": max_mem_clock,
                        "Driver Version": driver_version,
                        "Used By": self._get_gpu_usage(i),
                    }
                )

        except Exception as e:
            self.logger.error(f"Error getting GPU info: {e}")
            gpu_info.append(
                {
                    "GPU Index": "Error",
                    "Name": f"Failed to get GPU info: {str(e)}",
                    "Total Memory (GB)": "",
                    "Compute Capability": "",
                    "Max GPU Clock (MHz)": "",
                    "Max Memory Clock (MHz)": "",
                    "Driver Version": "",
                    "Used By": "",
                }
            )

        return pd.DataFrame(gpu_info)

    def _get_gpu_usage(self, gpu_index: int) -> str:
        """Determine what the GPU is used for based on configuration"""
        usage = []
        if gpu_index in self.vlm_gpus:
            usage.append("VLM")
        if gpu_index in self.llm_gpus:
            usage.append("LLM")
        return ", ".join(usage) if usage else "Not Used"

    def cleanup_resources(self):
        """Clean up any active resources"""
        with self.cleanup_lock:
            # Stop GPU monitoring
            if self.gpu_monitoring_active:
                self.stop_gpu_monitoring()

            # Clean up any tracked resources
            for resource in self.active_resources:
                try:
                    if isinstance(resource, str) and resource.startswith("stream_"):
                        # Clean up live stream
                        stream_id = resource.replace("stream_", "")
                        self.make_api_call(f"/live-stream/{stream_id}", method="DELETE")
                    elif isinstance(resource, str) and resource.startswith("file_"):
                        # Clean up uploaded file
                        file_id = resource.replace("file_", "")
                        self.make_api_call(f"/files/{file_id}", method="DELETE")
                except requests.exceptions.HTTPError as e:
                    # Ignore 400/404 errors - resource may already be deleted
                    if e.response is not None and e.response.status_code in [400, 404]:
                        self.logger.debug(f"Resource {resource} already deleted or not found")
                    else:
                        self.logger.error(f"Error cleaning up resource {resource}: {e}")
                except Exception as e:
                    self.logger.error(f"Error cleaning up resource {resource}: {e}")

            self.active_resources.clear()
            self.logger.debug("Resource cleanup completed")

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.cleanup_resources()
